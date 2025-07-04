from openpyxl import Workbook

wb= Workbook()
#wb.active
ws = wb.create_sheet() #새로운 sheet 기본 이름으로 생성
ws.title="MySheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "3CB371" #rgb값으로 색상 변경


# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet") #주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) #2번째 index에서 sheet생성

#시트별 작업
new_ws = wb["NewSheet"] #Dict형태로 sheet에 접근
print(wb.sheetnames) # 모든 sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("sample.xlsx")